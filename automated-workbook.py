
# Files & Directories 

# You can iterate over all the spreadsheets in a directory, open them and process them.

# Python program can automate thousands of excel spreadsheets in under a second. This can be built in less than 30 minutes.

# Openpyxl - package for working with excel spreadsheets.

# Procedures:-

# 1. Import openpyxl    2. Load Workbook    3. Iterating over rows     4. Fixing prices  5. Selecting values to add BarChart 6. Saving Workbook 

# Automated Workbook (code)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference 

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
         min_row=2, 
         max_row=sheet.max_row,
         min_col=4,
         max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')

    
# Automated Workbook Function

import openpyxl as xl
from openpyxl.chart import BarChart, Reference 

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
            min_row=2, 
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

# We’ve created a bar chart by calling openpyxl.chart.BarChart(). 
